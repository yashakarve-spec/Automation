import os
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter


def convert_file_to_excel():

    # ==========================================================
    # GET INPUT FILE
    # ==========================================================

    input_folder = input(
        "Enter input folder path: "
    ).strip().replace('"',"")

    if not os.path.isdir(input_folder):

        print("\nERROR: Folder not found. ")
        print("Path entered: ", input_folder)
        return
    files = sorted([ 
        f for f in os.listdir(input_folder)
        if re.fullmatch(r"\d{8}", f)
    ])

    if not files:

        print("\nNo YYYYMMDD files found. ")
        return
    
    for filename in files:
        input_file = os.path.join(
            input_folder,
            filename
        )

        print("\nProcessing:", filename)

        # ==========================================================
        # CREATE OUTPUT FOLDER
        # ==========================================================

        parent_folder = os.path.dirname(input_file)

        output_folder = os.path.join(parent_folder, "output")

        os.makedirs(output_folder, exist_ok=True)

        # ==========================================================
        # OUTPUT EXCEL FILE NAME
        # ==========================================================

        base_name = os.path.basename(input_file)

        excel_file = os.path.join(
            output_folder,
            base_name + ".xlsx"
        )

        # ==========================================================
        # READ INPUT FILE
        # ==========================================================

        try:
            with open(
                input_file,
                "r",
                encoding="utf-8",
                errors="ignore"
            ) as file:

                content = file.read()

        except Exception as e:
            print("Unable to read file.")
            print(e)
            return

        # ==========================================================
        # SPLIT RECORDS
        # ==========================================================

        records = [
            record.strip()
            for record in re.split(r"\r?\n", content)
            if record.strip()
        ]

        all_rows = []
        all_columns = []

        # ==========================================================
        # PARSE EACH RECORD
        # ==========================================================

        for record in records:

            parts = record.split(";", 3)

            row = {}

            if len(parts) > 0:
                row["Date_Time"] = parts[0]

            if len(parts) > 1:
                row["Event"] = parts[1]

            if len(parts) > 2:
                row["Job_ID"] = parts[2]

            if len(parts) > 3:

                remaining = parts[3].replace("\n", " ")

                matches = re.findall(
                    r'([A-Za-z0-9_.]+)=([^\s]+)',
                    remaining
                )

                for key, value in matches:
                    row[key] = value

            all_rows.append(row)

            for key in row:
                if key not in all_columns:
                    all_columns.append(key) 

        # ==========================================================
        # COLUMN ORDER
        # ==========================================================

        fixed_columns = [
            "Date_Time",
            "Event",
            "Job_ID"
        ]

        dynamic_columns = [
            col for col in all_columns
            if col not in fixed_columns
        ]

        columns = fixed_columns + dynamic_columns

        # ==========================================================
        # CREATE EXCEL
        # ==========================================================

        wb = Workbook()

        ws = wb.active
        ws.title = "PBS_Data"

        # ==========================================================
        # HEADER
        # ==========================================================

        for col_num, column_name in enumerate(columns, start=1):

            cell = ws.cell(
                row=1,
                column=col_num,
                value=column_name
            )

            cell.font = Font(bold=True)

        # ==========================================================
        # DATA
        # ==========================================================

        for row_num, row_data in enumerate(all_rows, start=2):

            for col_num, column_name in enumerate(columns, start=1):

                ws.cell(
                    row=row_num,
                    column=col_num,
                    value=row_data.get(column_name, "")
                )

        # ==========================================================
        # ADD BORDERS
        # ==========================================================

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in ws.iter_rows():

            for cell in row:

                cell.border = thin_border

        # ==========================================================
        # AUTO-FIT COLUMNS
        # ==========================================================

        for column_cells in ws.columns:

            max_length = 0

            for cell in column_cells:

                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except:
                    pass

            adjusted_width = min(max_length + 2, 60)

            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = adjusted_width

        # ==========================================================
        # SAVE FILE
        # ==========================================================

        try:

            wb.save(excel_file)

            print("\nSUCCESS")
            print("Excel file created:")
            print(excel_file)

        except Exception as e:

            print("Error saving Excel file")
            print(e)


if __name__ == "__main__":
    convert_file_to_excel()