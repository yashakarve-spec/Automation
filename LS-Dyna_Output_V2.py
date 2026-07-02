import os
import re
import math
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ==========================================================
# EXTRACT TIME VS AXIAL FORCE
# ==========================================================

def extract_all_beam_forces(elout_file):

    beam_data = {}

    with open(elout_file, "r", errors="ignore") as file:

        lines = file.readlines()

    current_time = None

    total_lines = len(lines)

    for i in range(total_lines):

        line = lines[i]

        # ======================================================
        # 1. EXTRACT TIME
        # ======================================================
        time_match = re.search(
            r'\(\s*at time\s*([+-]?\d+\.\d+E[+-]?\d+)\s*\)',
            line,
            re.IGNORECASE
        )

        if time_match:

            current_time = float(
                time_match.group(1)
            )

        # ======================================================
        # 2. DETECT BEAM/TRUSS BLOCK
        # ======================================================
        beam_match = re.search(
            r'beam/truss\s*#\s*=\s*(\d+)',
            line,
            re.IGNORECASE
        )

        if beam_match:

            beam_id = int(
                beam_match.group(1)
            )

            # initialize beam if not exists
            if beam_id not in beam_data:

                beam_data[beam_id] = []

            # ==================================================
            # 3. SEARCH RESULTANT BLOCK (LOOK AHEAD)
            # ==================================================

            for j in range(i + 1, total_lines):

                next_line = lines[j]

                # stop if next beam starts
                if "beam/truss" in next_line.lower():
                    break

                # find resultants block
                if "resultants" in next_line.lower():

                    # ==========================================
                    # 4. READ AXIAL FORCE VALUE
                    # ==========================================

                    for k in range(j + 1, total_lines):

                        value_line = lines[k].strip()

                        if not value_line:
                            continue

                        values = value_line.split()

                        try:

                            axial_force = float(values[0])

                            shear_s = float(values[1])

                            shear_t = float(values[2])

                            shear_resultant = math.sqrt(
                                shear_s**2 +
                                shear_t**2
                            )

                            if current_time is not None:

                                beam_data[beam_id].append(
                                    (
                                        current_time,
                                        axial_force,
                                        shear_s,
                                        shear_t,
                                        shear_resultant
                                    )
                                )

                            break

                        except (ValueError, IndexError):

                            continue

                    break

    return beam_data

# ==========================================================
# CREATE EXCEL
# ==========================================================

def create_excel(
    all_beams,
    resultant_beams,
    output_excel
):

    wb = Workbook()

    ws = wb.active
    ws.title = "Beam Forces"

    dashboard = wb.create_sheet("Dashboard")

    # ==========================================
    # DASHBOARD TABLE
    # ==========================================

    dashboard.cell(row=1, column=1, value="Bolt")
    dashboard.cell(row=1, column=2, value="Max Axial")
    dashboard.cell(row=1, column=3, value="Time @ Max Axial")
    dashboard.cell(row=1, column=4, value="Max Shear")
    dashboard.cell(row=1, column=5, value="Time @ Max Shear")

    for col in range(1, 6):

        dashboard.cell(
            row=1,
            column=col
        ).font = Font(
            name="Cambria",
            size=12,
            bold=True
        )

    dash_row = 2

    for beam_id, data in all_beams.items():

        if not str(beam_id).startswith("Resultant"):
            continue

        results = calculate_results(data)

        dashboard.cell(
            row=dash_row,
            column=1,
            value=str(beam_id)
        )

        dashboard.cell(
            row=dash_row,
            column=2,
            value=results["axial_max"][1]
        )

        dashboard.cell(
            row=dash_row,
            column=3,
            value=results["axial_max"][0]
        )

        dashboard.cell(
            row=dash_row,
            column=4,
            value=results["shear_max"][2]
        )

        dashboard.cell(
            row=dash_row,
            column=5,
            value=results["shear_max"][0]
        )

        dash_row += 1    

    start_col = 1

    for beam_id, data in all_beams.items():

        if not data:
            continue

        results = calculate_results(data)

        row = 1
        col = start_col

        # ==========================================
        # BEAM HEADING
        # ==========================================

        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + 2
        )

        heading_cell = ws.cell(
            row=row,
            column=col,
            value=f"Beam {beam_id}"
        )

        heading_cell.font = Font(
            name="Cambria",
            size=12,
            bold=True
        )

        heading_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        row += 1

        # ==========================================
        # TABLE HEADER
        # ==========================================

        time_header = ws.cell(
            row=row,
            column=col,
            value="Time (s)"
        )

        time_header.font = Font(
            name="Cambria",
            bold=True
        )

        time_header.alignment = Alignment(
            horizontal="center"
        )


        force_header = ws.cell(
            row=row,
            column=col + 1,
            value="Axial Force"
        )

        force_header.font = Font(
            name="Cambria",
            bold=True
        )

        force_header.alignment = Alignment(
            horizontal="center"
        )


        shear_header = ws.cell(
            row=row,
            column=col + 2,
            value="Shear Resultant"
        )

        shear_header.font = Font(
            name="Cambria",
            bold=True
        )

        shear_header.alignment = Alignment(
            horizontal="center"
        ) 

        row += 1

        # ==========================================
        # FORCE DATA
        # ==========================================

        for row_data in data:

            time_value = row_data[0]
            axial_force = row_data[1]
            shear_force = row_data[-1]

            ws.cell(
                row=row,
                column=col,
                value=time_value
            )

            ws.cell(
                row=row,
                column=col + 1,
                value=axial_force
            )

            ws.cell(
                row=row,
                column=col + 2,
                value=shear_force
            )

            row += 1

        # ==========================================
        # MAXIMUM FORCE
        # ==========================================

        ws.cell(
            row=row,
            column=col,
            value="Maximum Axial Force"
        )

        ws.cell(
            row=row,
            column=col + 1,
            value=results["axial_max"][1]
        )

        row += 1

        ws.cell(
            row=row,
            column=col,
            value="Time @ Max Axial"
        )

        ws.cell(
            row=row,
            column=col + 1,
            value=results["axial_max"][0]
        )

        row += 1

        ws.cell(
            row=row,
            column=col,
            value="Maximum Shear Force"
        )

        ws.cell(
            row=row,
            column=col + 1,
            value=results["shear_max"][2]
        )

        row += 1

        ws.cell(
            row=row,
            column=col,
            value="Time @ Max Shear"
        )

        ws.cell(
            row=row,
            column=col + 1,
            value=results["shear_max"][0]
        )

        # Leave 2 blank columns between tables
        start_col += 5

    # ==========================================
    # BORDERS
    # ==========================================

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row_cells in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column):

        for cell in row_cells:

            if cell.value is not None:

                cell.border = thin_border

    # ==========================================
    # COLUMN WIDTHS
    # ==========================================

    for col_num in range(1, ws.max_column + 1):

        ws.column_dimensions[
            get_column_letter(col_num)
        ].width = 25

    for row_cells in ws.iter_rows():

        for cell in row_cells:

            if cell.value is not None:

                cell.font = Font(
                    name="Cambria",
                    size=11,
                    bold=cell.font.bold
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
    # ==========================================
    # DASHBOARD CHARTS
    # ==========================================

    graph_start_row = 20
    chart_position_row = 2

    for bolt_name, bolt_data in resultant_beams.items():

        if not str(bolt_name).startswith("Resultant"):
            continue

        current_row = graph_start_row

        dashboard.cell(current_row, 1, "Time")
        dashboard.cell(current_row, 2, "Axial")
        dashboard.cell(current_row, 3, "Shear")

        current_row += 1

        for row_data in bolt_data:

            dashboard.cell(current_row, 1, row_data[0])
            dashboard.cell(current_row, 2, row_data[1])
            dashboard.cell(current_row, 3, row_data[2])

            current_row += 1

        # --------------------------------------
        # Axial Chart
        # --------------------------------------

        axial_chart = LineChart()
        axial_chart.title = f"{bolt_name} Axial"

        data = Reference(
            dashboard,
            min_col=2,
            min_row=graph_start_row,
            max_row=current_row - 1
        )

        categories = Reference(
            dashboard,
            min_col=1,
            min_row=graph_start_row + 1,
            max_row=current_row - 1
        )

        axial_chart.add_data(data, titles_from_data=True)
        axial_chart.set_categories(categories)

        dashboard.add_chart(
            axial_chart,
            f"G{chart_position_row}"
        )

        # --------------------------------------
        # Shear Chart
        # --------------------------------------

        shear_chart = LineChart()
        shear_chart.title = f"{bolt_name} Shear"

        data = Reference(
            dashboard,
            min_col=3,
            min_row=graph_start_row,
            max_row=current_row - 1
        )

        shear_chart.add_data(data, titles_from_data=True)
        shear_chart.set_categories(categories)

        dashboard.add_chart(
            shear_chart,
            f"O{chart_position_row}"
        )

        graph_start_row = current_row + 5
        chart_position_row += 15

    # ==========================================
    # HIDE GRAPH DATA
    # ==========================================

#    for r in range(20, current_row + 1):
#        dashboard.row_dimensions[r].hidden = True

    # ==========================================
    # DASHBOARD FORMATTING
    # ==========================================

    for col_num in range(1, 6):

        dashboard.column_dimensions[
            get_column_letter(col_num)
        ].width = 25

    for row_cells in dashboard.iter_rows(
            min_row=1,
            max_row=dash_row - 1,
            min_col=1,
            max_col=5):

        for cell in row_cells:

            if cell.value is not None:

                cell.border = thin_border

                cell.font = Font(
                    name="Cambria",
                    size=11,
                    bold=(cell.row == 1)
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    wb.save(output_excel)


# ==========================================================
# CALCULATE RESULTS
# ==========================================================

def calculate_results(data):

    axial_max = max(
        data,
        key=lambda x: abs(x[1])
    )

    shear_max = max(
        data,
        key=lambda x: abs(x[-1])
    )

    return {
        "axial_max": axial_max,
        "shear_max": shear_max
    }

# ==========================================================
# RESULTANT CALCULATION
# ==========================================================

def calculate_resultant_beams(all_beams):

    resultant_beams = {}

    beam_ids = sorted(all_beams.keys())

    for i in range(0, len(beam_ids) - 1, 2):

        beam1 = beam_ids[i]
        beam2 = beam_ids[i + 1]

        data1 = all_beams[beam1]
        data2 = all_beams[beam2]

        resultant_data = []
        
        print(data1[0])

        for (
                time1,
                axial1,
                shear_s1,
                shear_t1,
                shear_res1
            ), (
                time2,
                axial2,
                shear_s2,
                shear_t2,
                shear_res2
            ) in zip(
                data1,
                data2):

            bolt_axial = math.sqrt(
                axial1**2 +
                axial2**2
            )

            bolt_shear = math.sqrt(
                shear_res1**2 +
                shear_res2**2
            )

            resultant_data.append(
                (
                    time1,
                    bolt_axial,
                    bolt_shear
                )
            )

        resultant_beams[
            f"Resultant {beam1}-{beam2}"
        ] = resultant_data

    return resultant_beams

# ==========================================================
# MAIN
# ==========================================================

def main():

    print()
    print("===================================")
    print("LS-DYNA ELout Beam Force Extractor")
    print("===================================")
    print()

    elout_file = input(
        "Enter ELout file path : "
    ).strip()

    print()
    print("Reading entire ELout file and extracting all beams...")
    print()

    # ==========================================================
    # CREATE OUTPUT FOLDER
    # ==========================================================

    output_folder = r"Z:\09.1_User CAE_Data\22_Yash_Intern\From_Lalit\01_Beam_forces\Output"

    os.makedirs(
        output_folder,
        exist_ok=True
    )

    # Step 1: extract everything in one scan                
    all_beams = extract_all_beam_forces(elout_file)

    resultant_beams = calculate_resultant_beams(
    all_beams
    )

    if not all_beams:
        print("No beam data found in file")
        return

    base_name = os.path.splitext(
        os.path.basename(elout_file)
    )[0]

    excel_file = os.path.join(
        output_folder,
        f"{base_name}_Beam_Summary.xlsx"
    )

    all_beams.update(resultant_beams)

    create_excel(
        all_beams,
        resultant_beams,
        excel_file
    )
    
    print()
    print("Excel summary created successfully")
    print(os.path.abspath(excel_file))

if __name__ == "__main__":
    main()