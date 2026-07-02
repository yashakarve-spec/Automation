import os
from datetime import datetime, time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart,
    LineChart,
    Reference
)

def find_column(headers, possible_names):
    """
    Finds a column by checking multiple possible names.
    Returns the matching header name or None.
    """

    normalized = {
        str(h).strip().lower(): h
        for h in headers
        if h is not None
    }

    for name in possible_names:
        if name.lower() in normalized:
            return normalized[name.lower()]

    return None


def walltime_to_hours(value):
    """
    Converts walltime to decimal hours.
    Supports:
        HH:MM:SS
        H:MM:SS
        Excel time objects
    """

    if value is None:
        return 0.0

    if isinstance(value, time):
        return (
            value.hour
            + value.minute / 60
            + value.second / 3600
        )

    value = str(value).strip()

    if not value:
        return 0.0

    try:
        parts = value.split(":")

        if len(parts) == 3:
            h, m, s = map(int, parts)

            return (
                h
                + m / 60
                + s / 3600
            )

    except:
        pass

    return 0.0


def create_hpc_utilization_report():

    # ==========================================================
    # USER INPUT
    # ==========================================================

    input_folder = input(
        "Enter folder containing weekly Excel files: "
    ).strip().replace('"', '')

    if not os.path.isdir(input_folder):

        print("\nERROR: Folder not found")
        return

    # ==========================================================
    # FIND INPUT FILES
    # ==========================================================

    files = sorted([
        f for f in os.listdir(input_folder)
        if (
            f.endswith(".xlsx")
            and not f.startswith("~$")
            and len(os.path.splitext(f)[0]) == 8
            and os.path.splitext(f)[0].isdigit()
        )
    ])

    if not files:

        print("\nNo YYYYMMDD Excel files found.")
        return

    # ==========================================================
    # CONSTANTS
    # ==========================================================

    WEEKLY_CAPACITY = 144 * 24 * 7

    summary_rows = []

    application_totals = {}

    # ==========================================================
    # PROCESS FILES
    # ==========================================================

    for i in range(0, len(files), 7):

        weekly_files = files[i:i + 7]

        weekly_usage = 0.0

        for filename in weekly_files:

            filepath = os.path.join(
                input_folder,
                filename
            )

            print(f"Processing {filename}")

            try:

                wb = load_workbook(
                    filepath,
                    data_only=True
                )

            except Exception as e:

                print(f"Unable to open {filename}")
                print(e)
                continue

            # ==========================================================
            # FIND PBS_DATA SHEET
            # ==========================================================

            if "PBS_Data" in wb.sheetnames:

                ws = wb["PBS_Data"]

            else:

                ws = wb[wb.sheetnames[0]]

            # ==========================================================
            # READ HEADERS
            # ==========================================================

            headers = []

            for cell in ws[1]:
                headers.append(cell.value)

            # ==========================================================
            # FLEXIBLE COLUMN DETECTION
            # ==========================================================

            event_col = find_column(
                headers,
                [
                    "Event",
                    "event"
                ]
            )

            ncpus_col = find_column(
                headers,
                [
                    "resources_used.ncpus",
                    "resources_used_ncpus",
                    "ncpus",
                    "Resource_List.ncpus"
                ]
            )

            walltime_col = find_column(
                headers,
                [
                    "resources_used.walltime",
                    "resources_used_walltime",
                    "walltime"
                ]
            )

            software_col = find_column(
                headers,
                [
                    "Resource_List.software",
                    "resource_list.software",
                    "software"
                ]
            )

            if (
                event_col is None
                or ncpus_col is None
                or walltime_col is None
                or software_col is None
            ):

                print(
                    f"Required columns not found in {filename}"
                )
                continue

            # ==========================================================
            # HEADER INDEX LOOKUP
            # ==========================================================

            header_index = {
                str(cell.value): idx
                for idx, cell in enumerate(ws[1], start=1)
                if cell.value is not None
            }

            event_idx = header_index[event_col]
            ncpus_idx = header_index[ncpus_col]
            walltime_idx = header_index[walltime_col]
            software_idx = header_index[software_col]

            # ==========================================================
            # CALCULATE WEEKLY USAGE
            # ==========================================================

            for row in ws.iter_rows(
                min_row=2,
                values_only=True
            ):

                event = row[event_idx - 1]

                if str(event).strip() != "E":
                    continue

                try:

                    ncpus = float(
                        row[ncpus_idx - 1]
                    )

                except:

                    ncpus = 0

                walltime_hours = walltime_to_hours(
                    row[walltime_idx - 1]
                )

                core_hours = (
                    ncpus * walltime_hours
                )

                weekly_usage += core_hours

                software = str(
                    row[software_idx - 1]
                ).strip()

                if not software:
                    software = "Unknown"

                application_totals[software] = (
                    application_totals.get(
                        software,
                        0
                    )
                    + core_hours
                )

        # ==========================================================
        # DATE FROM LAST FILE OF THE WEEK
        # ==========================================================

        file_date = datetime.strptime(
            os.path.splitext(weekly_files[-1])[0],
            "%Y%m%d"
        )

        as_on_date = file_date.strftime(
            "%d-%b"
        )

        usage_percent = (
            weekly_usage
            / WEEKLY_CAPACITY
            * 100
        )

        summary_rows.append(
            [
                as_on_date,
                round(weekly_usage, 2),
                WEEKLY_CAPACITY,
                round(usage_percent, 2),
                100.0
            ]
        )

    # ==========================================================
    # CREATE OUTPUT WORKBOOK
    # ==========================================================

    output_wb = Workbook()

    ws = output_wb.active

    ws.title = "Weekly_Utilization"

    app_ws = output_wb.create_sheet(
        "Application_Utilization"
    )

    headers = [
        "As on Date",
        "Weekly Usage (Hrs)",
        "Weekly Capacity (Hrs)",
        "Weekly Usage (%)",
        "Weekly Capacity (%)"
    ]

    # ==========================================================
    # HEADER
    # ==========================================================

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=1,
            column=col_num,
            value=header
        )

        cell.font = Font(bold=True)
        cell.border = thin_border

    # ==========================================================
    # DATA
    # ==========================================================

    for row_num, row_data in enumerate(
        summary_rows,
        start=2
    ):

        for col_num, value in enumerate(
            row_data,
            start=1
        ):

            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=value
            )

            cell.border = thin_border

    # ==========================================================
    # AUTO WIDTH
    # ==========================================================

    for column_cells in ws.columns:

        max_length = 0

        for cell in column_cells:

            try:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except:
                pass

        ws.column_dimensions[
            get_column_letter(
                column_cells[0].column
            )
        ].width = min(
            max_length + 3,
            30
        )

    ws.freeze_panes = "A2"

    # ==========================================================
    # APPLICATION UTILIZATION SHEET
    # ==========================================================

    app_headers = [
        "Application",
        "Hours Consumed",
        "Usage (%)"
    ]

    for col_num, header in enumerate(
        app_headers,
        start=1
    ):

        cell = app_ws.cell(
            row=1,
            column=col_num,
            value=header
        )

        cell.font = Font(bold=True)
        cell.border = thin_border

    total_hours = sum(
        application_totals.values()
    )

    sorted_apps = sorted(
        application_totals.items(),
        key=lambda x: x[1],
        reverse=True
    )

    app_row = 2

    for app, hours in sorted_apps:

        usage_pct = (
            hours / total_hours * 100
        ) if total_hours else 0

        app_ws.cell(
            app_row,
            1,
            app
        )

        app_ws.cell(
            app_row,
            2,
            round(hours, 2)
        )

        app_ws.cell(
            app_row,
            3,
            round(usage_pct, 2)
        )

        for col in range(1, 4):

            app_ws.cell(
                app_row,
                col
            ).border = thin_border

        app_row += 1

    app_ws.cell(
        app_row,
        1,
        "Total"
    )

    app_ws.cell(
        app_row,
        2,
        round(total_hours, 2)
    )

    app_ws.cell(
        app_row,
        3,
        100
    )

    for col in range(1, 4):

        app_ws.cell(
            app_row,
            col
        ).border = thin_border

    last_app_row = app_row

    for column_cells in app_ws.columns:

        max_length = 0

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        app_ws.column_dimensions[
            get_column_letter(
                column_cells[0].column
            )
        ].width = min(
            max_length + 3,
            30
        )

    # ==========================================================
    # CHART CATEGORY REFERENCE
    # ==========================================================

    cats = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=len(summary_rows) + 1
    )

    # ==========================================================
    # CHART 1 - HOURS
    # ==========================================================

    bar_chart = BarChart()

    bar_chart.type = "col"
    bar_chart.style = 10

    bar_chart.title = (
        "Eka HPC Utilization Tracker (Hours)"
    )

    bar_chart.y_axis.title = (
        "Weekly Usage (Hrs)"
    )

    bar_chart.x_axis.title = (
        "As on Date"
    )

    usage_data = Reference(
        ws,
        min_col=2,
        max_col=2,
        min_row=1,
        max_row=len(summary_rows) + 1
    )

    bar_chart.add_data(
        usage_data,
        titles_from_data=True
    )

    bar_chart.set_categories(cats)

    line_chart = LineChart()

    capacity_data = Reference(
        ws,
        min_col=3,
        max_col=3,
        min_row=1,
        max_row=len(summary_rows) + 1
    )

    line_chart.add_data(
        capacity_data,
        titles_from_data=True
    )

    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"

    bar_chart += line_chart

    bar_chart.height = 12
    bar_chart.width = 24

    ws.add_chart(
        bar_chart,
        "G2"
    )

    # ==========================================================
    # CHART 2 - UTILIZATION %
    # ==========================================================

    bar_chart2 = BarChart()

    bar_chart2.type = "col"
    bar_chart2.style = 10

    bar_chart2.title = (
        "Eka HPC Utilization Tracker (%)"
    )

    bar_chart2.y_axis.title = (
        "Weekly Usage (%)"
    )

    bar_chart2.x_axis.title = (
        "As on Date"
    )

    usage_pct = Reference(
        ws,
        min_col=4,
        max_col=4,
        min_row=1,
        max_row=len(summary_rows) + 1
    )

    bar_chart2.add_data(
        usage_pct,
        titles_from_data=True
    )

    bar_chart2.set_categories(cats)

    line_chart2 = LineChart()

    capacity_pct = Reference(
        ws,
        min_col=5,
        max_col=5,
        min_row=1,
        max_row=len(summary_rows) + 1
    )

    line_chart2.add_data(
        capacity_pct,
        titles_from_data=True
    )

    line_chart2.y_axis.axId = 300
    line_chart2.y_axis.crosses = "max"

    bar_chart2 += line_chart2

    bar_chart2.height = 12
    bar_chart2.width = 24

    ws.add_chart(
        bar_chart2,
        "G24"
    )

    # ==========================================================
    # APPLICATION HOURS CHART
    # ==========================================================

    app_chart1 = BarChart()

    app_chart1.type = "col"
    app_chart1.style = 10

    app_chart1.title = (
        "Application-wise Usage (Hours)"
    )

    app_chart1.y_axis.title = (
        "Hours Consumed"
    )

    app_chart1.x_axis.title = (
        "Application"
    )

    data = Reference(
        app_ws,
        min_col=2,
        max_col=2,
        min_row=1,
        max_row=last_app_row - 1
    )

    cats = Reference(
        app_ws,
        min_col=1,
        min_row=2,
        max_row=last_app_row - 1
    )

    app_chart1.add_data(
        data,
        titles_from_data=True
    )

    app_chart1.set_categories(cats)

    app_chart1.height = 12
    app_chart1.width = 20

    app_ws.add_chart(
        app_chart1,
        "F2"
    )

    # ==========================================================
    # APPLICATION % CHART
    # ==========================================================

    app_chart2 = BarChart()

    app_chart2.type = "col"
    app_chart2.style = 10

    app_chart2.title = (
        "Application-wise Usage (%)"
    )

    app_chart2.y_axis.title = (
        "% Utilization"
    )

    app_chart2.x_axis.title = (
        "Application"
    )

    data = Reference(
        app_ws,
        min_col=3,
        max_col=3,
        min_row=1,
        max_row=last_app_row - 1
    )

    app_chart2.add_data(
        data,
        titles_from_data=True
    )

    app_chart2.set_categories(cats)

    app_chart2.height = 12
    app_chart2.width = 20

    app_ws.add_chart(
        app_chart2,
        "F22"
    )

    # ==========================================================
    # SAVE
    # ==========================================================

    output_file = os.path.join(
        input_folder,
        "HPC_Weekly_Utilization.xlsx"
    )

    output_wb.save(output_file)

    print("\nSUCCESS")
    print("Output file:")
    print(output_file)


if __name__ == "__main__":
    create_hpc_utilization_report()